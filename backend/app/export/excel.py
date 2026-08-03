import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelService:
    @staticmethod
    def generate_applications_excel(applications: list) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Internship Applications"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")

        headers = ["ID", "Full Name", "Email", "Phone", "College", "Degree", "Year of Study", "Skills", "Duration", "Status", "Applied At"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 26

        for row_idx, app in enumerate(applications, 2):
            skills_str = ", ".join(app.skills) if isinstance(app.skills, list) else str(app.skills)
            created_str = app.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(app, 'created_at') and app.created_at else ""

            row_data = [
                app.id,
                app.full_name,
                app.email,
                app.phone,
                app.college,
                app.degree,
                app.year_of_study,
                skills_str,
                app.duration,
                app.status.upper(),
                created_str
            ]
            ws.append(row_data)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx in [1, 7, 9, 10, 11]:
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_payments_excel(payments: list) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payments History"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")
        alignment_right = Alignment(horizontal="right", vertical="center")

        headers = ["Payment ID", "Order ID", "Student Email", "Amount (INR)", "Status", "Registration ID", "Timestamp"]
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border

        ws.row_dimensions[1].height = 26

        for row_idx, pmt in enumerate(payments, 2):
            created_str = pmt.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(pmt, 'created_at') and pmt.created_at else ""

            row_data = [
                pmt.payment_id or "N/A",
                pmt.order_id,
                pmt.student_email,
                pmt.amount_inr,
                pmt.status.upper(),
                pmt.registration_id or "N/A",
                created_str
            ]
            ws.append(row_data)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx == 4:
                    cell.alignment = alignment_right
                    cell.number_format = '#,##0'
                elif col_idx in [1, 2, 5, 6, 7]:
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

excel_service = ExcelService()
